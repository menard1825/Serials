import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_serials_excel(client_name, order_number, client_po, hardware_data, software_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Serial Numbers & Licenses"

    # Safari Micro branding colors
    header_color = "004785"
    separator_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Header formatting
    ws.merge_cells("A1:D3")
    ws["A4"] = "Safari Micro - Serial Number & License Report"
    ws["A4"].font = Font(size=16, bold=True, color=header_color)
    ws["A4"].alignment = Alignment(horizontal="left")

    # Customer message
    ws["A6"] = "Dear Valued Customer," 
    ws["A6"].font = Font(size=12, italic=True)
    ws["A7"] = "Please find below the serial numbers and/or license keys for your order. If you need any assistance, don't hesitate to reach out."
    ws["A7"].font = Font(size=12)

    # Order details
    ws["A9"], ws["B9"] = "Client Name:", client_name
    ws["A10"], ws["B10"] = "Order Number:", order_number
    ws["A11"], ws["B11"] = "Client PO:", client_po

    row_index = 13
    
    # Hardware Section
    if hardware_data:
        ws[row_index][0].value = "Hardware - Serial Numbers"
        ws[row_index][0].font = Font(bold=True, size=14)
        row_index += 1

        headers = ["Product", "Model Number", "Serial Number"]
        ws.append(headers)
        for col in range(1, 4):
            cell = ws.cell(row=row_index, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.border = border_style
        row_index += 1

        for product, model, serial_numbers in hardware_data:
            if row_index > 14:
                for col in range(1, 4):
                    ws.cell(row=row_index, column=col).fill = separator_fill
                row_index += 1
            for serial in serial_numbers:
                ws[f"A{row_index}"] = product
                ws[f"B{row_index}"] = model
                ws[f"C{row_index}"] = serial
                for col in range(1, 4):
                    ws.cell(row=row_index, column=col).border = border_style
                row_index += 1
    
    # Software Section
    if software_data:
        row_index += 2
        ws[row_index][0].value = "Software - License Keys"
        ws[row_index][0].font = Font(bold=True, size=14)
        row_index += 1

        headers = ["Software Name", "License Type", "License Key"]
        ws.append(headers)
        for col in range(1, 4):
            cell = ws.cell(row=row_index, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.border = border_style
        row_index += 1

        for software, license_type, license_keys in software_data:
            for key in license_keys:
                ws[f"A{row_index}"] = software
                ws[f"B{row_index}"] = license_type
                ws[f"C{row_index}"] = key
                for col in range(1, 4):
                    ws.cell(row=row_index, column=col).border = border_style
                row_index += 1
    
    # Adjust column width
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45

    # Save to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Streamlit UI
st.set_page_config(page_title="Safari Micro Serial & License Tool", page_icon="📄", layout="centered")
st.image("https://safarimicro.com/wp-content/uploads/2022/01/SafariMicro-Color-with-Solid-Icon-Copy.png", width=300)

st.title("Safari Micro Serial & License Tool")

client_name = st.text_input("Client Name", key="client_name")
order_number = st.text_input("Order Number", key="order_number")
client_po = st.text_input("Client PO", key="client_po")

st.write("Select the type of data you want to enter:")
data_type = st.radio("Data Type", ["Hardware - Serial Numbers", "Software - License Keys", "Both"], key="data_type")

hardware_data = []
software_data = []

if data_type in ["Hardware - Serial Numbers", "Both"]:
    st.write("Enter Hardware Details:")
    num_products = st.number_input("Number of Hardware Items", min_value=1, step=1, key="num_products")
    for i in range(num_products):
        product_name = st.text_input(f"Product {i+1} Name", key=f"product_name_{i}")
        model_number = st.text_input(f"Product {i+1} Model Number", key=f"model_number_{i}")
        serial_input = st.text_area(f"Enter Serial Numbers for Product {i+1} (comma-separated)", "", key=f"serial_input_{i}")
        serials = [s.strip() for s in serial_input.split(",") if s.strip()]
        hardware_data.append((product_name, model_number, serials))

if data_type in ["Software - License Keys", "Both"]:
    st.write("Enter Software Details:")
    num_software = st.number_input("Number of Software Items", min_value=1, step=1, key="num_software")
    for i in range(num_software):
        software_name = st.text_input(f"Software {i+1} Name", key=f"software_name_{i}")
        license_type = st.text_input(f"License Type for {software_name}", key=f"license_type_{i}")
        key_input = st.text_area(f"Enter License Keys for {software_name} (comma-separated)", "", key=f"key_input_{i}")
        keys = [k.strip() for k in key_input.split(",") if k.strip()]
        software_data.append((software_name, license_type, keys))

if st.button("Generate Excel File", key="generate_excel"):
    excel_data = create_serials_excel(client_name, order_number, client_po, hardware_data, software_data)
    file_name = f"SafariMicro_{client_name}_{order_number}.xlsx".replace(" ", "_")
    st.download_button(label="Download Excel File", data=excel_data, file_name=file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
